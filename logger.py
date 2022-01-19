try:
    import sys
    import webbrowser
    import logging
    import os
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print('One of the libraries failed to import.')
    sys.exit()

'''grab current files in the same directory as the script'''

files = os.listdir('.')

''' set up variables and create a list of all current excel files '''

months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12', 'Jan': '01',
          'Feb': '02', 'Mar': '03', 'Apr': '04', 'Jun': '06', 'Jul': '07',
          'Aug': '08', 'Sep': '09', 'Oc': '10', 'Nov': '11', 'Dec': '12'}
years = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']
directories = [item for item in files if '.xlsx' in item]


''' set up how the log file will be recorded and we call this right before our main function'''


def setter(bool):
    if not bool:
        logging.basicConfig(filename='log.log', filemode='w', format='%(asctime)s - %(levelname)s %(message)s',
                            datefmt='%H:%M', encoding='utf-8', level=logging.INFO)
    else:
        logging.basicConfig(filename='log.log', filemode='a', format='%(asctime)s - %(levelname)s %(message)s',
                            datefmt='%H:%M', encoding='utf-8', level=logging.INFO)


''' This function shows the users which files they can scan and lets them choose.
If they dont enter a correct option it loops through the
while loop again and it returns the chosen file name minus the '.xlsx'. 
If directories is empty that means no .xlsx files were found so we raise an error to exit out'''


def workb():
    picked = False
    while not picked:
        try:
            if len(directories) == 0:
                raise FileNotFoundError
            for num,item in enumerate(directories, 1):
                print(f'{num} - {item}')
            which = input("Please pick which file you would like to get "
                          "data from using it's corresponding number above: ")
            if which.lower() == 'exit':
                sys.exit()
            if which.isdigit():
                if int(which) == 0:
                    print("0 isn't an option please pick from the options provided")
                    continue
                if int(which) <= (len(directories)):
                    picked = True
                else:
                    print('That option was not a choice please try again')
                    continue
            else:
                print('That option was not a choice, please try again')
        except FileNotFoundError:
            print('File was not found. Please make sure the script and the files you '
                  'want to scan are in the same directory.')
            sys.exit()
    answer = directories[int(which) - 1][:-5]
    directories.pop(int(which) - 1)
    return answer

''' this is a function to see if we should go more data . '''


def would():
    ask = input("Would you like get data from another file (Y/N): ")
    if ask.lower() == 'exit':
        sys.exit()
    if ask.lower() == "yes" or ask.lower() == "y":
        loginfo('--' * 12)
        return True
    if ask.lower() == "no" or ask.lower() == "n":
        return False
    else:
        print("Looks like you gave a bad value that wasn't expected, please answer yes or no. :)")
        would()


def openlog():
    ask = input("Would you like to me to open the log file for you? (Y/N): ")
    if ask.lower() == 'exit':
        sys.exit()
    if ask.lower() == "yes" or ask.lower() == "y":
        webbrowser.open('log.log')
        sys.exit()
    if ask.lower() == "no" or ask.lower() == "n":
        print("Okay have a good day.")
        sys.exit()
    else:
        print("Looks like you gave a bad value that wasn't expected, please answer yes or no. :)")
        openlog()


''' created this to help with logging all the info '''


def loginfo(*info):
    for item in info:
        logging.info(item)


setter(False)


def main():

    '''gets the current workbook using current variable and adding the .xlsx extention back and then uses the current
    variable to get the month and date, then we split the month and date into a 2018/01 format which matches datetime
    formats of yyyy/mm/dd without the /dd that excel is returning. '''

    current = workb()
    wb = load_workbook(current + '.xlsx')
    try:
        mandd = list(set([item for item in current.split("_")[::-1] if item.capitalize() in months or item in years]))
        if len(mandd) <= 1 or len(mandd) > 2:
            raise ValueError
    except ValueError:
        print(f'Looks like something might be misspelled or something extra is'
              f' in your filename as we only got this - {mandd}')
        sys.exit()
    if not mandd[0].isdigit():
        mandd = mandd[::-1]
    mand = f'{mandd[0]}-{months[mandd[1].capitalize()]}'
    ws = wb['Summary Rolling MoM']

    ''' set the row to none and then we use a try statement to find the row that the correct date is on,
    if no date is found exits the program with a value error '''
    row = None
    try:
        for item in ws['A']:
            if mand in str(item.value):
                row = item.row
        if row is None:
            raise ValueError
    except ValueError:
        print('Was unable to find the correct row. Make sure it follows standard formatting procedures.')
        sys.exit()





    ''' then we use the row we just got to get the info required and then using the loginfo function
    earlier to log the info. '''

    vals = list(ws.iter_rows(min_row=row, max_row=row, values_only=True))
    vals = [item for item in vals[0][1:] if item is not None]
    loginfo(f"Month of {mandd[1].capitalize()}, {mandd[0]} ", f"Calls Offered: {vals[0]}",
            f"Abandon after 30s: {round(vals[1]*100,2)}%",
            f"FCR: {vals[2]*100}0%", f"DSAT: {vals[3]*100}0%", f"CSAT: {vals[4]*100}0%")


    ''' we swap to the second sheet and then go through the first row to find the month we need. If the 
    datetime format match is not found we then look for the just plain jane month match, and if that fails 
    we exit the program with a value error. we use the Z variable to keep track of how far over the column is 
    that we need to go through once we find it. '''

    ws = wb["VOC Rolling MoM"]
    z = 0
    found = False
    try:
        for item in ws['1']:
            z += 1
            if mand in str(item.value):
                found = True
                break
        if not found:
            z=0
            for item in ws['1']:
                z+= 1
                if str(item.value).capitalize() in months:
                    found = True
                    break
        if not found:
            raise ValueError
    except ValueError:
        print('Was unable to find the correct column. Make sure it follows standard formatting procedures.')
        sys.exit()

    ''' we then create a new list with each item in the column
    (we get the column by using the tool get_column_letter) 
    then we log the info using the loginfo function '''

    nums = [item.value for item in ws[get_column_letter(z)][1:] if item.value is not None and int(item.value) != 0][1:]
    if nums[0] >= 200:
        promote = f"Promoters: {nums[0]}, which is good!"
    else:
        promote = f"Promoters: {nums[0]}, which is bad!"
    if nums[1] >= 100:
        passives = f"Passives: {nums[1]}, which is good!"
    else:
        passives = f"Passives: {nums[1]}, which is bad!"
    if nums[2] > 100:
        dec = f"Dectractors: {nums[2]}, which is good!"
    else:
        dec = f"Dectractors: {nums[2]}, which is bad!"
    loginfo(f"In Net Promoter Score - Base Size = {nums[0]}",promote,passives,dec)

    ''' now we ask if you want to get data from multiple files and then if yes we run this again, if not we ask if they 
    want to pull up the log'''

    if would():
        setter(True)
        main()
    else:
        openlog()


if __name__ == "__main__":
    main()



